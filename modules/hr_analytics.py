"""
modules/hr_analytics.py  ·  FinOx Suite
=========================================
HR Analytics: Workforce Intelligence & Risk Detection.

CACHING CHANGES (v2.0)
-----------------------
• _make_template_xlsx() — wrapped with @st.cache_data. This module-level
  function builds a fully styled openpyxl workbook in memory and returns it as
  bytes. It takes ZERO arguments, so its cache key is always the same constant —
  it runs once and is served from cache on every subsequent render. Before
  caching, it was rebuilding the workbook from scratch on every Streamlit rerun
  just to populate the download button.

  The return type (bytes) is trivially serializable by Streamlit's pickle-based
  cache. No UnhashableTypeError risk.

  All other methods are intentionally NOT cached:
  - _load_section():    Contains st.file_uploader() and st.checkbox() UI calls.
                        Calling UI widgets inside cached functions raises a
                        ScriptRunContext error and would crash the app.
  - _coerce():          Mutates a DataFrame copy. Not a pure function.
  - All _tab_*():       Contain st.plotly_chart(), st.dataframe() UI calls.
  - _render_insights(): Calls self._insight_box() which makes a live AI API call.
"""
from __future__ import annotations

import io
from typing import Any

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from core.base_module import BaseModule
from utils.formatters import fmt, read_file, safe_div

_CUR    = "\u09f3"
_NAVY   = "#0B1F3A"
_TEAL   = "#00B4A6"
_AMBER  = "#F5A623"
_RED    = "#E84855"
_GREEN  = "#2ECC71"
_PURPLE = "#9B59B6"

_PAL = [_TEAL, _NAVY, _AMBER, _RED, _GREEN, _PURPLE, "#3498DB", "#E67E22"]

_BASE = dict(
    template      = "plotly_white",
    paper_bgcolor = "rgba(0,0,0,0)",
    plot_bgcolor  = "rgba(0,0,0,0)",
    font          = dict(family="'Segoe UI',Arial,sans-serif", size=12, color="#2D3748"),
    hoverlabel    = dict(bgcolor=_NAVY, font_color="#FFF", font_size=12),
    margin        = dict(t=52, b=22, l=12, r=12),
)

_SAMPLE = pd.DataFrame({
    "Name":              ["Alice", "Bob", "Carol", "Dave", "Eve", "Frank",
                          "Grace", "Hank", "Iris", "Jack", "Karen", "Leo"],
    "Department":        ["Sales", "Sales", "Marketing", "Operations", "Operations",
                          "Technology", "Technology", "HR", "Finance", "Finance",
                          "Sales", "Technology"],
    "Years_Experience":  [5, 8, 4, 3, 6, 10, 7, 2, 9, 5, 3, 12],
    "Salary":            [80_000, 95_000, 70_000, 60_000, 65_000, 110_000,
                          125_000, 55_000, 90_000, 85_000, 72_000, 140_000],
    "Performance_Score": [4.5, 4.8, 4.2, 4.0, 3.8, 4.9, 4.7, 3.5, 4.3, 4.1, 3.9, 5.0],
    "Absenteeism_Days":  [2, 1, 3, 5, 4, 1, 2, 8, 2, 3, 6, 1],
    "Training_Hours":    [40, 35, 30, 20, 25, 50, 45, 15, 38, 32, 18, 60],
})

_REQUIRED = ["Name", "Department", "Salary", "Performance_Score"]
_NUMERIC  = ["Salary", "Performance_Score", "Years_Experience",
             "Absenteeism_Days", "Training_Hours"]


@st.cache_data
def _make_template_xlsx() -> bytes:
    """
    Build and return a styled HR data Excel template as bytes.

    @st.cache_data: This function takes no arguments, so its cache key is a
    constant — it executes once per server process and is served from cache on
    every subsequent call. The openpyxl workbook creation is non-trivial and
    was previously rebuilt from scratch on every single Streamlit rerun just to
    populate the download button. The return type (bytes) is trivially
    serializable — no UnhashableTypeError risk.
    """
    buf = io.BytesIO()
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "HR Data"
        headers = ["Name", "Department", "Years_Experience", "Salary",
                   "Performance_Score", "Absenteeism_Days", "Training_Hours"]
        fill = PatternFill("solid", fgColor="00B4A6")
        bold = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
        ws.append(["Alice Johnson", "Sales", 5, 80000, 4.5, 2, 40])
        ws.append(["Bob Smith",     "Technology", 8, 110000, 4.8, 1, 50])
        wb.save(buf)
    except ImportError:
        buf.write(
            _SAMPLE[["Name", "Department", "Years_Experience", "Salary",
                     "Performance_Score", "Absenteeism_Days", "Training_Hours"]]
            .to_csv(index=False).encode()
        )
    return buf.getvalue()


class HRAnalyticsModule(BaseModule):
    PAGE_ICON  = "👨‍💼"
    PAGE_TITLE = "HR Analytics"

    def render(self) -> None:
        self._page_header(
            "👨‍💼 HR Analytics: Workforce Intelligence",
            "Upload your HR data (.csv or .xlsx) · or explore the built-in sample dataset",
        )
        df = self._load_section()
        if df is None or df.empty:
            return
        if not self._require_columns(df, _REQUIRED):
            return
        df = self._coerce(df)

        self._render_kpi_strip(df)
        st.markdown("---")

        tab1, tab2, tab3, tab4 = st.tabs([
            "🎻 Compensation", "📊 Performance", "🏢 Departments", "⚠️ Flight Risk",
        ])
        with tab1: self._tab_compensation(df)
        with tab2: self._tab_performance(df)
        with tab3: self._tab_departments(df)
        with tab4: self._tab_risk(df)

        st.markdown("---")
        self._render_insights(df)

    def _load_section(self) -> pd.DataFrame | None:
        c_up, c_info = st.columns([2, 1], gap="medium")
        with c_info:
            tpl  = _make_template_xlsx()
            ext  = ".xlsx" if tpl[:4] == b"PK\x03\x04" else ".csv"
            mime = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                if ext == ".xlsx" else "text/csv"
            )
            st.download_button(
                label="⬇️ Download Sample Template",
                data=tpl, file_name=f"hr_template{ext}", mime=mime,
            )
            use_sample = st.checkbox("Use built-in sample data", value=True, key="hr_use_sample")
        with c_up:
            uploaded = st.file_uploader(
                "Upload HR Data  (.csv, .xlsx, .xls)",
                type=["csv", "xlsx", "xls"], key="hr_upload",
                help="Required columns: " + ", ".join(_REQUIRED),
            )
        if uploaded is not None:
            try:
                df = read_file(uploaded)
                st.success(
                    f"✅ Loaded **{len(df):,}** rows · **{len(df.columns)}** columns "
                    f"from **{uploaded.name}**"
                )
                with st.expander("Preview uploaded data", expanded=False):
                    st.dataframe(df.head(10), use_container_width=True, hide_index=True)
                return df
            except Exception as exc:
                st.error(f"Could not read **{uploaded.name}**: {exc}")
        if use_sample:
            st.info("ℹ️ Showing built-in sample data (12 employees). Upload your own file for real analysis.")
            return _SAMPLE.copy()
        st.warning("Upload a file or enable sample data to begin.")
        return None

    def _coerce(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        for col in _NUMERIC:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        return df.dropna(subset=["Salary", "Performance_Score"])

    def _render_kpi_strip(self, df: pd.DataFrame) -> None:
        n        = len(df)
        payroll  = df["Salary"].sum()
        avg_sal  = df["Salary"].mean()
        avg_perf = df["Performance_Score"].mean()
        ann_rev  = self.price * self.qty * 12
        rev_emp  = safe_div(ann_rev, n)
        top_dept = (
            df.groupby("Department")["Performance_Score"].mean().idxmax()
            if "Department" in df.columns and n > 0 else "—"
        )
        c1, c2, c3, c4, c5, c6 = st.columns(6, gap="small")
        c1.metric("👥 Headcount",        str(n))
        c2.metric("💼 Annual Payroll",   fmt(payroll))
        c3.metric("💰 Avg Salary",       fmt(avg_sal))
        c4.metric("⭐ Avg Performance",  f"{avg_perf:.2f} / 5.0")
        c5.metric("📈 Revenue/Employee", fmt(rev_emp))
        c6.metric("🏆 Top Department",   top_dept)

    def _tab_compensation(self, df: pd.DataFrame) -> None:
        col1, col2 = st.columns([3, 2], gap="medium")
        with col1:
            st.markdown(f"<h5 style='color:{_NAVY}'>Salary Distribution by Department</h5>", unsafe_allow_html=True)
            hover = {c: True for c in ["Name", "Performance_Score"] if c in df.columns}
            fig = px.violin(
                df, x="Department", y="Salary", color="Department",
                color_discrete_sequence=_PAL, box=True, points="all", hover_data=hover,
            )
            fig.update_layout(**_BASE, height=380, showlegend=False,
                              xaxis_title="",
                              yaxis=dict(title="Annual Salary (৳)", tickprefix=_CUR, gridcolor="#EDF2F7"))
            st.plotly_chart(fig, use_container_width=True)
        with col2:
            st.markdown(f"<h5 style='color:{_NAVY}'>Payroll Share by Department</h5>", unsafe_allow_html=True)
            dept_pay = (
                df.groupby("Department")["Salary"].sum()
                .reset_index().rename(columns={"Salary": "Total"})
                .sort_values("Total", ascending=False)
            )
            fig2 = go.Figure(go.Pie(
                labels=dept_pay["Department"], values=dept_pay["Total"],
                hole=0.52, marker=dict(colors=_PAL, line=dict(color="#FFF", width=2)),
                textinfo="label+percent", textfont=dict(size=11),
                hovertemplate="<b>%{label}</b><br>৳%{value:,.0f}<br>%{percent}<extra></extra>",
            ))
            fig2.update_layout(**_BASE, height=260, showlegend=False)
            st.plotly_chart(fig2, use_container_width=True)
            med = df["Salary"].median()
            hi  = df["Salary"].max()
            lo  = df["Salary"].min()
            with st.container(border=True):
                top_name = df.loc[df["Salary"].idxmax(), "Name"] if "Name" in df.columns else "—"
                st.markdown(
                    f"**Median:** {fmt(med)}  \n"
                    f"**Range:** {fmt(lo)} – {fmt(hi)}  \n"
                    f"**Top earner:** {top_name}"
                )

    def _tab_performance(self, df: pd.DataFrame) -> None:
        if "Years_Experience" not in df.columns:
            self._info_box("Years_Experience column missing — performance scatter unavailable.")
            return
        fig = px.scatter(
            df, x="Years_Experience", y="Performance_Score",
            color="Department" if "Department" in df.columns else None,
            size="Salary", hover_name="Name" if "Name" in df.columns else None,
            color_discrete_sequence=_PAL, size_max=25,
            trendline="ols" if len(df) > 3 else None,
            title="Performance vs Experience (bubble size = Salary)",
        )
        fig.update_layout(**_BASE, height=380,
                          xaxis=dict(title="Years of Experience", gridcolor="#EDF2F7"),
                          yaxis=dict(title="Performance Score (1–5)", gridcolor="#EDF2F7"))
        st.plotly_chart(fig, use_container_width=True)
        if "Training_Hours" in df.columns:
            st.markdown("#### Training Hours vs Performance")
            fig2 = px.scatter(
                df, x="Training_Hours", y="Performance_Score",
                color="Department" if "Department" in df.columns else None,
                trendline="ols" if len(df) > 3 else None,
                title="Training Hours vs Performance Score",
                template="plotly_white",
            )
            st.plotly_chart(fig2, use_container_width=True)

    def _tab_departments(self, df: pd.DataFrame) -> None:
        if "Department" not in df.columns:
            self._info_box("Department column missing.")
            return
        dept = (
            df.groupby("Department")
            .agg(
                Headcount=("Name", "count"),
                Avg_Salary=("Salary", "mean"),
                Avg_Perf=("Performance_Score", "mean"),
                Total_Payroll=("Salary", "sum"),
            )
            .reset_index()
        )
        dept["Cost_Per_Perf"] = dept["Total_Payroll"] / dept["Avg_Perf"].replace(0, 1)
        dept_sorted = dept.sort_values("Cost_Per_Perf")
        fig = px.bar(
            dept_sorted, x="Cost_Per_Perf", y="Department", orientation="h",
            color="Avg_Perf", color_continuous_scale="RdYlGn",
            title="Payroll Efficiency (Total Payroll ÷ Avg Performance)",
            template="plotly_white",
        )
        fig.update_layout(xaxis=dict(tickprefix=_CUR))
        st.plotly_chart(fig, use_container_width=True)
        st.dataframe(
            dept.style.format({
                "Avg_Salary":    fmt,
                "Total_Payroll": fmt,
                "Avg_Perf":      "{:.2f}",
                "Cost_Per_Perf": fmt,
            }),
            use_container_width=True, hide_index=True,
        )

    def _tab_risk(self, df: pd.DataFrame) -> None:
        if "Absenteeism_Days" not in df.columns or "Years_Experience" not in df.columns:
            self._info_box("Absenteeism_Days and Years_Experience required for flight-risk analysis.")
            return
        df2 = df.copy()
        df2["Risk_Score"] = (
            (df2["Absenteeism_Days"] / df2["Absenteeism_Days"].max().clip(1)) * 50
            + ((5 - df2["Performance_Score"]) / 4) * 50
        )
        df2["Risk_Level"] = pd.cut(
            df2["Risk_Score"], bins=[0, 30, 60, 100],
            labels=["🟢 Low", "🟡 Medium", "🔴 High"]
        )
        fig = px.scatter(
            df2, x="Years_Experience", y="Salary",
            color="Risk_Level",
            color_discrete_map={"🟢 Low": _GREEN, "🟡 Medium": _AMBER, "🔴 High": _RED},
            size="Risk_Score", hover_name="Name" if "Name" in df2.columns else None,
            title="Flight Risk Matrix (size = Risk Score)",
            template="plotly_white",
        )
        st.plotly_chart(fig, use_container_width=True)

        high_risk = df2[df2["Risk_Level"] == "🔴 High"]
        if not high_risk.empty:
            st.markdown(f"**⚠️ {len(high_risk)} High-Risk Employees — Cost Exposure: {fmt(high_risk['Salary'].sum())}**")
            st.dataframe(
                high_risk[["Name", "Department", "Salary", "Performance_Score",
                            "Absenteeism_Days", "Risk_Score"]].style.format(
                    {"Salary": fmt, "Risk_Score": "{:.1f}"}
                ),
                use_container_width=True, hide_index=True,
            )

    def _render_insights(self, df: pd.DataFrame) -> None:
        payroll  = df["Salary"].sum()
        avg_perf = df["Performance_Score"].mean()
        ann_rev  = self.price * self.qty * 12
        self._insight_box(
            what=(
                f"Headcount: {len(df)}, annual payroll: {fmt(payroll)}, "
                f"avg performance: {avg_perf:.2f}/5.0, "
                f"revenue per employee: {fmt(safe_div(ann_rev, len(df)))}."
            ),
            recommendation=(
                "If revenue-per-employee is below payroll cost per head, "
                "prioritise upskilling the bottom quartile performers. "
                "Address high-risk employees before they leave — replacement costs "
                "typically run 50–200% of annual salary."
            ),
            context_data={
                "Headcount":             len(df),
                "Annual Payroll":        fmt(payroll),
                "Avg Performance":       f"{avg_perf:.2f}/5",
                "Revenue Per Employee":  fmt(safe_div(ann_rev, len(df))),
                "Payroll/Revenue Ratio": f"{safe_div(payroll, ann_rev):.1%}",
            },
        )